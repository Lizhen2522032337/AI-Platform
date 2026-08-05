"""LangGraph Agent 的目录校验、安全边界和文件 Tool 测试。"""

import asyncio
import io
import zipfile
from types import SimpleNamespace

import pytest
from app.agent import graph, planners
from app.agent.catalog import CatalogQuery, QueryCatalog
from app.agent.database_tool import _prepare_sql
from app.agent.db2_tool import Db2ToolError, _assert_read_only_sql
from app.agent.dynamic_sql_tool import (
    DynamicSqlToolError,
    execute_dynamic_sql,
    validate_dynamic_sql,
)
from app.agent.report_tools import create_report_files
from app.agent.types import AgentPlan, DynamicSqlQuery
from app.config.settings import Settings
from app.dify import KnowledgeResult
from openpyxl import load_workbook


def settings(**overrides) -> Settings:
    values = {
        "minio_root_user": "test-user",
        "minio_root_password": "test-password",
        "deepseek_api_key": "test-deepseek-key",
        "qwen_api_key": "test-qwen-key",
        "qwen_base_url": "https://example.invalid/v1",
        "postgres_password": "test-postgres-password",
    }
    values.update(overrides)
    return Settings(**values)


@pytest.mark.parametrize(
    "sql",
    [
        "UPDATE PROD.STATE SET VALUE = 1",
        "SELECT * FROM PROD.STATE; DELETE FROM PROD.STATE",
        "CALL PROD.RESET_STATE()",
    ],
)
def test_db2_tool_rejects_non_read_only_sql(sql: str) -> None:
    with pytest.raises(Db2ToolError):
        _assert_read_only_sql(sql)


def test_db2_tool_accepts_select_and_with() -> None:
    _assert_read_only_sql("SELECT ID FROM PROD.STATE FETCH FIRST 10 ROWS ONLY")
    _assert_read_only_sql("WITH X AS (SELECT ID FROM PROD.STATE) SELECT ID FROM X")


def test_report_file_tool_creates_office_pdf_json_and_dynamic_sql_csv(monkeypatch) -> None:
    stored = []

    def fake_save_file(object_key, payload, content_type):
        stored.append((object_key, payload, content_type))
        return {"objectKey": object_key, "contentType": content_type, "size": len(payload)}

    monkeypatch.setattr("app.agent.report_tools.save_file", fake_save_file)
    artifacts = create_report_files(
        9,
        "停机分析",
        "## 摘要\n发现压力异常。",
        [
            {
                "tool": "dynamic_sql",
                "status": "ok",
                "columns": ["username", "role_code"],
                "rows": [{"username": "admin", "role_code": "admin"}],
            }
        ],
        settings(report_files_enabled=True),
    )
    assert len(artifacts) == 6
    stored_by_key = {object_key: payload for object_key, payload, _ in stored}
    assert stored_by_key["tasks/9/report.md"].startswith(b"# ")
    assert zipfile.is_zipfile(io.BytesIO(stored_by_key["tasks/9/report.docx"]))
    assert stored_by_key["tasks/9/report.pdf"].startswith(b"%PDF")
    assert zipfile.is_zipfile(io.BytesIO(stored_by_key["tasks/9/report.xlsx"]))
    workbook = load_workbook(io.BytesIO(stored_by_key["tasks/9/report.xlsx"]))
    assert workbook.sheetnames == ["报告", "platform_data"]
    assert workbook["platform_data"]["A2"].value == "admin"
    assert b"username" in stored_by_key["tasks/9/platform_data.csv"]
    assert b"observations" in stored_by_key["tasks/9/evidence.json"]


def test_langgraph_routes_report_and_builds_evidence(monkeypatch) -> None:
    trace_steps = []

    async def fake_create_plan(
        intent,
        prompt,
        provider,
        messages,
        catalog,
        database_type,
        allow_dynamic_sql=False,
    ):
        assert intent == "report_generation"
        assert database_type == "postgresql"
        assert allow_dynamic_sql is False
        return AgentPlan(
            intent=intent,
            objective=prompt,
            knowledge_query="生产日报指标",
            report_required=True,
            report_title="生产日报",
        )

    async def fake_retrieve(query):
        assert query == "生产日报指标"
        return KnowledgeResult(True, "[知识库 1]\n日报口径", 1)

    monkeypatch.setattr(graph, "create_plan", fake_create_plan)
    monkeypatch.setattr(graph, "retrieve_knowledge", fake_retrieve)
    monkeypatch.setattr(graph, "load_query_catalog", lambda: QueryCatalog())
    monkeypatch.setattr(graph, "get_settings", lambda: settings())

    prepared = asyncio.run(
        graph.prepare_agent(
            12,
            "生成今天的生产日报",
            "qwen",
            [{"role": "user", "content": "生成今天的生产日报"}],
            trace_callback=trace_steps.append,
        )
    )
    assert prepared.intent == "report_generation"
    assert prepared.plan.report_required is True
    assert "日报口径" in prepared.context
    assert any(step["id"] == "dify_knowledge" for step in trace_steps)
    assert all("日报口径" not in str(step) for step in trace_steps)


@pytest.mark.parametrize(
    ("prompt", "expected"),
    [
        ("整理当前所有用户", "platform_data_query"),
        ("查询禁用账号", "platform_data_query"),
        ("统计各个角色有多少用户", "platform_data_query"),
        ("整理当前所有任务，导出 Excel", "platform_data_query"),
        ("生成今天的生产日报", "report_generation"),
        ("分析设备停机原因", "incident_analysis"),
    ],
)
def test_supervisor_routes_platform_user_queries(prompt: str, expected: str) -> None:
    assert planners.choose_intent(prompt) == expected


def test_catalog_query_model_allows_parameterized_select() -> None:
    query = CatalogQuery(
        id="production_window",
        description="查询时间窗内生产记录",
        sql="SELECT ID FROM PROD.RECORD WHERE CREATED_AT >= ?",
    )
    assert query.id == "production_window"
    assert query.supports("db2") is True
    assert query.supports("postgresql") is False


def test_catalog_selects_sql_for_requested_database_dialect() -> None:
    query = CatalogQuery(
        id="production_window",
        description="查询时间窗内生产记录",
        sql={
            "postgresql": "SELECT id FROM record WHERE created_at >= :start_time LIMIT 10",
            "db2": "SELECT id FROM record WHERE created_at >= :start_time FETCH FIRST 10 ROWS ONLY",
        },
        parameters=[
            {
                "name": "start_time",
                "description": "开始时间",
                "type": "string",
            }
        ],
    )
    assert "LIMIT 10" in (query.sql_for("postgresql") or "")
    assert "FETCH FIRST 10" in (query.sql_for("db2") or "")

    postgresql_sql, postgresql_values = _prepare_sql(
        query.sql_for("postgresql") or "",
        "postgresql",
        query.parameters,
        {"start_time": "2026-07-31T00:00:00+08:00"},
    )
    db2_sql, db2_values = _prepare_sql(
        query.sql_for("db2") or "",
        "db2",
        query.parameters,
        {"start_time": "2026-07-31T00:00:00+08:00"},
    )
    assert "%s" in postgresql_sql
    assert "?" in db2_sql
    assert postgresql_values == db2_values


def test_dynamic_sql_accepts_explicit_platform_user_join() -> None:
    normalized, tables = validate_dynamic_sql(
        """
        SELECT u.id, u.username, u.display_name, r.code AS role_code,
               u.is_active, u.last_login_at, u.created_at
        FROM public.app_users AS u
        JOIN public.auth_roles AS r ON r.id = u.role_id
        ORDER BY u.created_at DESC
        """
    )

    assert "password_hash" not in normalized
    assert tables == ["public.app_users", "public.auth_roles"]


def test_dynamic_sql_accepts_safe_aggregate_and_cte() -> None:
    aggregate, _ = validate_dynamic_sql(
        "SELECT COUNT(*) AS user_count FROM public.app_users WHERE is_active IS TRUE"
    )
    cte, _ = validate_dynamic_sql(
        """
        WITH active AS (
          SELECT id, username FROM public.app_users WHERE is_active IS TRUE
        )
        SELECT username FROM active ORDER BY username
        """
    )

    assert "COUNT(*)" in aggregate
    assert "WITH active" in cte


def test_dynamic_sql_accepts_admin_task_list() -> None:
    normalized, tables = validate_dynamic_sql(
        """
        SELECT t.id AS task_id, t.prompt, t.status, t.model_provider,
               u.username AS created_by_username, t.created_at
        FROM public.ai_tasks AS t
        LEFT JOIN public.app_users AS u ON u.id = t.created_by
        ORDER BY t.created_at DESC
        """
    )

    assert "answer" not in normalized
    assert tables == ["public.ai_tasks", "public.app_users"]


@pytest.mark.parametrize(
    "sql",
    [
        "SELECT * FROM public.app_users",
        "SELECT password_hash FROM public.app_users",
        "SELECT answer FROM public.ai_tasks",
        "SELECT id FROM app_users",
        "SELECT pg_read_file('/etc/passwd') FROM public.app_users",
        "SELECT pg_sleep(10) FROM public.app_users",
        "SELECT id INTO TEMP copied_users FROM public.app_users",
        "UPDATE public.app_users SET is_active = FALSE",
        "SELECT id FROM public.app_users; DELETE FROM public.app_users",
        "SELECT tableoid, id AS tableoid FROM public.app_users",
        """SELECT id FROM public.app_users
           WHERE EXISTS (SELECT id AS tableoid FROM public.auth_roles)
           ORDER BY tableoid""",
    ],
)
def test_dynamic_sql_rejects_unsafe_statements(sql: str) -> None:
    with pytest.raises(DynamicSqlToolError):
        validate_dynamic_sql(sql)


def test_dynamic_sql_executes_in_read_only_transaction_and_limits_rows(
    monkeypatch,
) -> None:
    calls = []

    class FakeCursor:
        description = None

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def execute(self, statement, parameters=None):
            calls.append((statement, parameters))
            if "dynamic_result" in statement:
                self.description = [SimpleNamespace(name="username")]

        def fetchall(self):
            return [("admin",), ("second",)]

    class FakeConnection:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def cursor(self):
            return FakeCursor()

    monkeypatch.setattr(
        "app.agent.dynamic_sql_tool.psycopg.connect",
        lambda **_kwargs: FakeConnection(),
    )
    result = execute_dynamic_sql(
        "SELECT username FROM public.app_users ORDER BY username",
        settings(dynamic_sql_max_rows=1),
    )

    assert calls[0][0] == "SET TRANSACTION READ ONLY"
    assert calls[-1][1] == (2,)
    assert result["rows"] == [{"username": "admin"}]
    assert result["truncated"] is True


def test_planner_removes_dynamic_sql_without_admin_permission(monkeypatch) -> None:
    async def fake_complete_json(_provider, _system_prompt, _user_prompt):
        return {
            "intent": "platform_data_query",
            "objective": "整理所有用户",
            "knowledge_query": "平台用户",
            "hypotheses": [],
            "queries": [],
            "dynamic_query": {
                "purpose": "查询用户",
                "sql": "SELECT id, username FROM public.app_users",
            },
            "report_required": False,
            "report_title": "用户清单",
            "notify": False,
        }

    monkeypatch.setattr(planners, "complete_json", fake_complete_json)
    plan = asyncio.run(
        planners.create_plan(
            "platform_data_query",
            "整理所有用户",
            "deepseek",
            [{"role": "user", "content": "整理所有用户"}],
            QueryCatalog(),
            "postgresql",
            settings(),
            allow_dynamic_sql=False,
        )
    )

    assert plan.dynamic_query is None

    admin_plan = asyncio.run(
        planners.create_plan(
            "platform_data_query",
            "整理所有用户",
            "deepseek",
            [{"role": "user", "content": "整理所有用户"}],
            QueryCatalog(),
            "postgresql",
            settings(),
            allow_dynamic_sql=True,
        )
    )
    assert admin_plan.dynamic_query is not None


def test_platform_planner_supplies_safe_query_when_model_omits_it(monkeypatch) -> None:
    async def fake_complete_json(_provider, system_prompt, _user_prompt):
        assert "必须生成 dynamic_query" in system_prompt
        return {
            "intent": "platform_data_query",
            "objective": "整理当前所有用户",
            "knowledge_query": "平台用户",
            "hypotheses": [],
            "queries": [],
            "dynamic_query": None,
            "report_required": False,
            "report_title": "平台用户查询",
            "notify": False,
        }

    monkeypatch.setattr(planners, "complete_json", fake_complete_json)
    plan = asyncio.run(
        planners.create_plan(
            "platform_data_query",
            "整理当前所有用户",
            "deepseek",
            [{"role": "user", "content": "整理当前所有用户"}],
            QueryCatalog(),
            "postgresql",
            settings(),
            allow_dynamic_sql=True,
        )
    )

    assert plan.dynamic_query is not None
    normalized, tables = validate_dynamic_sql(plan.dynamic_query.sql)
    assert "password_hash" not in normalized
    assert tables == ["public.app_users", "public.auth_roles"]


def test_platform_task_planner_supplies_safe_query_when_model_omits_it(
    monkeypatch,
) -> None:
    async def fake_complete_json(_provider, _system_prompt, _user_prompt):
        return {
            "intent": "platform_data_query",
            "objective": "整理当前所有任务并导出 Excel",
            "knowledge_query": "平台任务",
            "hypotheses": [],
            "queries": [],
            "dynamic_query": None,
            "report_required": False,
            "report_title": "平台任务清单",
            "notify": False,
        }

    monkeypatch.setattr(planners, "complete_json", fake_complete_json)
    plan = asyncio.run(
        planners.create_plan(
            "platform_data_query",
            "整理当前所有任务并导出 Excel",
            "deepseek",
            [{"role": "user", "content": "整理当前所有任务并导出 Excel"}],
            QueryCatalog(),
            "postgresql",
            settings(),
            allow_dynamic_sql=True,
        )
    )

    assert plan.report_required is True
    assert plan.dynamic_query is not None
    normalized, tables = validate_dynamic_sql(plan.dynamic_query.sql)
    assert "answer" not in normalized
    assert tables == ["public.ai_tasks", "public.app_users"]


def test_platform_planner_forces_report_files_when_user_requests_excel(monkeypatch) -> None:
    async def fake_complete_json(_provider, _system_prompt, _user_prompt):
        return {
            "intent": "platform_data_query",
            "objective": "整理当前所有用户并导出 Excel",
            "knowledge_query": "平台用户",
            "hypotheses": [],
            "queries": [],
            "dynamic_query": None,
            "report_required": False,
            "report_title": "平台用户清单",
            "notify": False,
        }

    monkeypatch.setattr(planners, "complete_json", fake_complete_json)
    plan = asyncio.run(
        planners.create_plan(
            "platform_data_query",
            "整理当前所有用户并导出 Excel",
            "deepseek",
            [{"role": "user", "content": "整理当前所有用户并导出 Excel"}],
            QueryCatalog(),
            "postgresql",
            settings(),
            allow_dynamic_sql=True,
        )
    )

    assert plan.report_required is True
    assert plan.dynamic_query is not None


def test_platform_planner_replaces_unsafe_model_sql(monkeypatch) -> None:
    async def fake_complete_json(_provider, _system_prompt, _user_prompt):
        return {
            "intent": "platform_data_query",
            "objective": "整理当前所有用户",
            "knowledge_query": "平台用户",
            "hypotheses": [],
            "queries": [],
            "dynamic_query": {
                "purpose": "查询全部用户字段",
                "sql": "SELECT * FROM public.app_users",
            },
            "report_required": False,
            "report_title": "平台用户查询",
            "notify": False,
        }

    monkeypatch.setattr(planners, "complete_json", fake_complete_json)
    plan = asyncio.run(
        planners.create_plan(
            "platform_data_query",
            "整理当前所有用户",
            "deepseek",
            [{"role": "user", "content": "整理当前所有用户"}],
            QueryCatalog(),
            "postgresql",
            settings(),
            allow_dynamic_sql=True,
        )
    )

    assert plan.dynamic_query is not None
    normalized, _ = validate_dynamic_sql(plan.dynamic_query.sql)
    assert "SELECT *" not in normalized
    assert "password_hash" not in normalized


def test_admin_platform_query_runs_dynamic_sql_through_full_graph(monkeypatch) -> None:
    traces = []

    async def fake_complete_json(_provider, _system_prompt, _user_prompt):
        return {
            "intent": "platform_data_query",
            "objective": "整理当前所有用户",
            "knowledge_query": "平台用户",
            "hypotheses": [],
            "queries": [],
            "dynamic_query": None,
            "report_required": False,
            "report_title": "平台用户查询",
            "notify": False,
        }

    monkeypatch.setattr(planners, "complete_json", fake_complete_json)
    monkeypatch.setattr(planners, "get_settings", lambda: settings())
    monkeypatch.setattr(graph, "get_settings", lambda: settings())
    monkeypatch.setattr(graph, "load_query_catalog", lambda: QueryCatalog())
    monkeypatch.setattr(
        graph,
        "execute_dynamic_sql",
        lambda _sql: {
            "tool": "dynamic_sql",
            "databaseType": "postgresql",
            "status": "ok",
            "queryFingerprint": "test",
            "tables": ["public.app_users", "public.auth_roles"],
            "columns": ["username", "role_code"],
            "rows": [{"username": "admin", "role_code": "admin"}],
            "truncated": False,
        },
    )

    prepared = asyncio.run(
        graph.prepare_agent(
            21,
            "整理当前所有用户",
            "deepseek",
            [{"role": "user", "content": "整理当前所有用户"}],
            database_type="postgresql",
            allow_dynamic_sql=True,
            trace_callback=traces.append,
        )
    )

    assert prepared.intent == "platform_data_query"
    assert prepared.plan.dynamic_query is not None
    assert prepared.observations[0]["rows"][0]["username"] == "admin"
    assert "优先使用 Markdown 表格" in prepared.context
    assert any(
        step["id"] == "supervisor" and "平台数据查询" in step.get("detail", "")
        for step in traces
    )
    assert any(
        step["id"] == "dynamic_sql_access" and step["status"] == "completed"
        for step in traces
    )
    assert any(
        step["id"] == "dynamic_sql" and step["status"] == "completed"
        for step in traces
    )
    assert all(step["id"] != "dify_knowledge" for step in traces)


def test_database_node_executes_admin_dynamic_sql(monkeypatch) -> None:
    traces = []
    monkeypatch.setattr(graph, "load_query_catalog", lambda: QueryCatalog())
    monkeypatch.setattr(
        graph,
        "execute_dynamic_sql",
        lambda sql: {
            "tool": "dynamic_sql",
            "status": "ok",
            "rows": [{"username": "admin"}],
            "columns": ["username"],
            "truncated": False,
            "queryFingerprint": "test",
        },
    )
    plan = AgentPlan(
        intent="incident_analysis",
        objective="整理所有用户",
        knowledge_query="平台用户",
        dynamic_query=DynamicSqlQuery(
            purpose="查询用户",
            sql="SELECT username FROM public.app_users",
        ),
    )

    result = asyncio.run(
        graph.database_tool_node(
            {
                "task_id": 18,
                "database_type": "postgresql",
                "allow_dynamic_sql": True,
                "plan": plan.model_dump(),
                "observations": [],
                "trace_callback": traces.append,
            }
        )
    )

    assert result["observations"][0]["tool"] == "dynamic_sql"
    assert any(step["id"] == "dynamic_sql" for step in traces)
